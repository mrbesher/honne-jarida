import openpyxl
import sys

if len(sys.argv) < 2:
    sys.exit("usage: migrate.py <bookkeeping.xlsx>")

wb = openpyxl.load_workbook(sys.argv[1], data_only=True)

def q(v):
    if v is None or v == "":
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"

def emit(date, t, cents, cat, sub, src, note):
    print(f"INSERT INTO transactions (date, type, amount_cents, category, subcategory, source, note) VALUES ({q(date)}, '{t}', {cents}, {q(cat)}, {q(sub)}, {q(src)}, {q(note)});")

print("BEGIN;")

for sheet in wb.sheetnames:
    if sheet in ("Summary", "Income"):
        continue
    for row in wb[sheet].iter_rows(min_row=2, values_only=True):
        date, expense, total, cat, sub, src, note = row[:7]
        if date is None or total is None:
            continue
        iso = date.strftime("%Y-%m-%d") if hasattr(date, "strftime") else str(date)
        cents = round(float(total) * 100)
        full_note = " | ".join(filter(None, [expense, note]))
        emit(iso, "expense", cents, cat, sub, src, full_note)

for row in wb["Income"].iter_rows(min_row=2, values_only=True):
    date, amount, source = row[:3]
    if date is None or amount is None:
        continue
    iso = date.strftime("%Y-%m-%d") if hasattr(date, "strftime") else str(date)
    cents = round(float(amount) * 100)
    emit(iso, "income", cents, None, None, None, source)

print("COMMIT;")
